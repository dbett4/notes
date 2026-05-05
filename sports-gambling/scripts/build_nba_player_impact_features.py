#!/usr/bin/env python3
"""Build NBA player-impact priors from the Jeremias/RAPM workbook.

This script creates feature inputs only. It does not emit betting picks.
"""

from __future__ import annotations

import argparse
import csv
import math
from pathlib import Path
from statistics import mean, pstdev
from typing import Any

from openpyxl import load_workbook

COUNT_COLUMNS = [
    "FG2_miss",
    "FG3_miss",
    "FG2_make_ua",
    "FG3_make_ua",
    "FG2_make_a",
    "FG3_make_a",
    "FT_miss",
    "FT_make",
    "live_to_bad_pass",
    "live_to_lost_ball",
    "pts",
    "shooting_foul",
    "fouled_3",
    "steal_bad_pass",
    "steal_lost_ball",
    "dead_to",
    "goaltend",
    "blocked",
    "And1s_made",
    "And1s_missed",
    "And1_fouls",
    "blocks_to_def",
    "blocks_to_off",
    "dreb_2",
    "dreb_3",
    "dreb_ft",
    "DEFLECTIONS",
    "CHARGES_DRAWN",
    "def_6ft_make",
    "def_6ft_miss",
    "def_6t10_make",
    "def_6t10_miss",
    "def_long2_make",
    "def_long2_miss",
    "def_3_make",
    "def_3_miss",
    "oreb",
    "dreb_contested_close",
    "dreb_uncontested_close",
    "dreb_contested_far",
    "dreb_uncontested_far",
    "assists",
]

BASE_COLUMNS = [
    "name",
    "possessions",
    "gp",
    "starts",
    "mp_approx",
    "seasons",
    "off_RAPM",
    "def_RAPM",
    "total_RAPM",
    "off_RAPM_SE",
    "def_RAPM_SE",
    "rapm_reliability",
    "possession_reliability",
    "season_reliability",
    "starter_reliability",
    "model_status",
]


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def per_100(value: float | None, possessions: float | None) -> float | None:
    if value is None or not possessions or possessions <= 0:
        return None
    return value / possessions * 100


def clamp(value: float, low: float = 0.0, high: float = 1.0) -> float:
    return max(low, min(high, value))


def z_scores(rows: list[dict[str, Any]], columns: list[str]) -> None:
    for column in columns:
        values = [row[column] for row in rows if isinstance(row.get(column), float)]
        if not values:
            continue
        mu = mean(values)
        sigma = pstdev(values) or 1.0
        for row in rows:
            value = row.get(column)
            row[f"{column}_z"] = None if value is None else (value - mu) / sigma


def rounded(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return round(value, 6)
    return value


def load_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows_iter)]

    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        raw = dict(zip(headers, values))
        possessions = parse_float(raw.get("possessions"))
        off_rapm = parse_float(raw.get("off_RAPM"))
        def_rapm = parse_float(raw.get("def_RAPM"))
        off_se = parse_float(raw.get("off_RAPM_SE"))
        def_se = parse_float(raw.get("def_RAPM_SE"))
        starts = parse_float(raw.get("starts"))
        gp = parse_float(raw.get("gp"))
        seasons = parse_float(raw.get("seasons"))

        row: dict[str, Any] = {
            "name": raw.get("name"),
            "possessions": possessions,
            "gp": gp,
            "starts": starts,
            "mp_approx": parse_float(raw.get("mp_approx")),
            "seasons": seasons,
            "off_RAPM": off_rapm,
            "def_RAPM": def_rapm,
            "total_RAPM": (off_rapm or 0.0) + (def_rapm or 0.0),
            "off_RAPM_SE": off_se,
            "def_RAPM_SE": def_se,
            "rapm_reliability": None,
            "possession_reliability": None,
            "season_reliability": None,
            "starter_reliability": None,
            "model_status": "experimental_player_prior_not_pick",
        }

        if off_se is not None and def_se is not None:
            row["rapm_reliability"] = 1 / (1 + ((off_se + def_se) / 2))
        if possessions is not None:
            row["possession_reliability"] = clamp(possessions / 100000)
        if seasons is not None:
            row["season_reliability"] = clamp(seasons / 10)
        if starts is not None and gp:
            row["starter_reliability"] = starts / gp

        for column in COUNT_COLUMNS:
            row[f"{column}_per_100"] = per_100(parse_float(raw.get(column)), possessions)

        rows.append(row)

    z_scores(rows, [f"{column}_per_100" for column in COUNT_COLUMNS])
    z_scores(rows, ["off_RAPM", "def_RAPM", "total_RAPM"])
    return rows


def write_rows(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    per_100_columns = [f"{column}_per_100" for column in COUNT_COLUMNS]
    z_columns = [f"{column}_z" for column in per_100_columns] + [
        "off_RAPM_z",
        "def_RAPM_z",
        "total_RAPM_z",
    ]
    headers = BASE_COLUMNS + per_100_columns + z_columns

    with path.open("w", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: rounded(row.get(header)) for header in headers})


def validate(path: Path, expected_rows_including_header: int = 1478) -> None:
    rows = list(csv.DictReader(path.open()))
    required = {
        "name",
        "possessions",
        "mp_approx",
        "seasons",
        "off_RAPM",
        "def_RAPM",
        "total_RAPM",
        "off_RAPM_SE",
        "def_RAPM_SE",
        "rapm_reliability",
        "possession_reliability",
        "season_reliability",
        "starter_reliability",
        "model_status",
    }
    missing = required - set(rows[0].keys())
    if missing:
        raise SystemExit(f"Missing required columns: {sorted(missing)}")
    actual = len(rows) + 1
    if actual != expected_rows_including_header:
        raise SystemExit(f"Expected {expected_rows_including_header} rows including header, got {actual}")
    if {row["model_status"] for row in rows} != {"experimental_player_prior_not_pick"}:
        raise SystemExit("Output must label every row as experimental player priors")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="data/raw/nba/jeremias-nba-metric.xlsx")
    parser.add_argument("--output", default="data/processed/nba/player_impact_features.csv")
    parser.add_argument("--validate", action="store_true")
    args = parser.parse_args()

    rows = load_rows(Path(args.input))
    write_rows(rows, Path(args.output))
    if args.validate:
        validate(Path(args.output))
    print(f"Wrote {len(rows) + 1} rows including header to {args.output}")


if __name__ == "__main__":
    main()
